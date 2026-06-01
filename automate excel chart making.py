"""Generated from Jupyter notebook: Automate Excel Chart Making with Python

Magics and shell lines are commented out. Run with a normal Python interpreter."""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

df: pd.DataFrame | None = None
quarterly_sales: pd.DataFrame | None = None
file_path: str | None = None
folder_path: str | None = None
regions: list | None = None


def section_2_loading_our_data() -> None:
    global df
    df = pd.read_excel(
        "https://github.com/datagy/pivot_table_pandas/raw/master/sample_pivot.xlsx",
        parse_dates=["Date"],
    )
    df.head()


def section_3_testing_pivot_tables() -> None:
    global quarterly_sales
    filtered = df[df["Region"] == "East"]
    quarterly_sales = pd.pivot_table(
        filtered,
        index=filtered["Date"].dt.quarter,
        columns="Type",
        values="Sales",
        aggfunc="sum",
    )
    print("Quarterly Sales Pivot Table:")
    quarterly_sales.head()


def section_04_creating_and_excel_workbook() -> None:
    global file_path
    file_path = "data/test.xlsx"
    quarterly_sales.to_excel(file_path, sheet_name="Quarterly Sales", startrow=3)


def section_05_loading_the_workbook() -> None:
    wb = load_workbook(file_path)
    sheet1 = wb["Quarterly Sales"]
    sheet1["A1"] = "Quarterly Sales"
    sheet1["A2"] = "datagy.io"
    sheet1["A4"] = "Quarter"
    sheet1["A1"].style = "Title"
    sheet1["A2"].style = "Headline 2"
    for i in range(5, 9):
        sheet1[f"B{i}"].style = "Currency"
        sheet1[f"C{i}"].style = "Currency"
        sheet1[f"D{i}"].style = "Currency"

    bar_chart = BarChart()
    data = Reference(sheet1, min_col=2, max_col=4, min_row=4, max_row=8)
    categories = Reference(sheet1, min_col=1, max_col=1, min_row=5, max_row=8)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    sheet1.add_chart(bar_chart, "F4")
    bar_chart.title = "Sales by Type"
    bar_chart.style = 3
    wb.save(filename=file_path)


def section_08_getting_region_names() -> None:
    global regions, folder_path, file_path, quarterly_sales
    regions = list(df["Region"].unique())
    folder_path = "data/test"
    for region in regions:
        filtered = df[df["Region"] == f"{region}"]
        quarterly_sales = pd.pivot_table(
            filtered,
            index=filtered["Date"].dt.quarter,
            columns="Type",
            values="Sales",
            aggfunc="sum",
        )
        file_path = f"{folder_path}{region}.xlsx"
        quarterly_sales.to_excel(file_path, sheet_name="Quarterly Sales", startrow=3)
        wb = load_workbook(file_path)
        sheet1 = wb["Quarterly Sales"]
        sheet1["A1"] = "Quarterly Sales"
        sheet1["A2"] = "datagy.io"
        sheet1["A4"] = "Quarter"
        sheet1["A1"].style = "Title"
        sheet1["A2"].style = "Headline 2"
        for i in range(5, 10):
            sheet1[f"B{i}"].style = "Currency"
            sheet1[f"C{i}"].style = "Currency"
            sheet1[f"D{i}"].style = "Currency"
        bar_chart = BarChart()
        data = Reference(sheet1, min_col=2, max_col=4, min_row=4, max_row=8)
        categories = Reference(sheet1, min_col=1, max_col=1, min_row=5, max_row=8)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        sheet1.add_chart(bar_chart, "F4")
        bar_chart.title = "Sales by Type"
        bar_chart.style = 3
        wb.save(file_path)


def main() -> None:
    section_2_loading_our_data()
    section_3_testing_pivot_tables()
    section_04_creating_and_excel_workbook()
    section_05_loading_the_workbook()
    section_08_getting_region_names()


if __name__ == "__main__":
    main()
